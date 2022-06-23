import logging
import string
import typing
from collections import Counter

import numpy as np
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA

from .. import QGRAIN_VERSION
from ..artificial._generator import ArtificialDataset
from ..emma import EMMAResult
from ..model import GrainSizeDataset, GrainSizeSample
from ..ssu import DISTRIBUTION_CLASS_MAP, SSUResult
from ..statistic._GRADISTAT import _get_all_scales, get_all_statistic, logarithmic
from ._use_excel import column_to_char, prepare_styles

SMALL_WIDTH = 12
MEDIAN_WIDTH = 24
LARGE_WIDTH = 48


def _check_dataset(dataset: GrainSizeDataset):
    if dataset is None:
        raise ValueError("The dataset is `None`.")
    elif not dataset.has_sample:
        raise ValueError("There is no sample in this dataset.")


def _write_readme_sheet(ws: Worksheet, text: str):
    ws.title = "README"
    full_text = \
        """
        This Excel file was generated by QGrain ({0}).

        Please cite:
        Liu, Y., Liu, X., Sun, Y., 2021. QGrain: An open-source and easy-to-use software for the comprehensive analysis of grain size distributions. Sedimentary Geology 423, 105980.
        DOI: 10.1016/j.sedgeo.2021.105980

        {1}

        """.format(QGRAIN_VERSION, text)
    lines_of_description = full_text.split("\n")
    for row, line in enumerate(lines_of_description):
        cell = ws.cell(row+1, 1, value=line)
        cell.style = "description"
    ws.column_dimensions[column_to_char(0)].width = 200


def _write_dataset_sheet(
        ws: Worksheet,
        dataset: GrainSizeDataset,
        progress_callback: typing.Callable = None):
    def write(row, col, value, style="normal_light"):
        cell = ws.cell(row+1, col+1, value=value)
        cell.style = style

    ws.title = "GSDs"
    write(0, 0, "Sample Name", style="header")
    ws.column_dimensions[column_to_char(0)].width = 24
    for col, value in enumerate(dataset.classes_μm, 1):
        write(0, col, value, style="header")
        ws.column_dimensions[column_to_char(col)].width = 10

    for i_sample, sample in enumerate(dataset.samples):
        row = i_sample + 1
        if row % 2 == 0:
            style = "normal_dark"
        else:
            style = "normal_light"
        write(row, 0, sample.name, style=style)
        for col, value in enumerate(sample.distribution, 1):
            write(row, col, value, style=style)
        if progress_callback is not None:
            progress = i_sample / dataset.n_samples
            progress_callback(progress)


def save_artificial_dataset(
        dataset: ArtificialDataset,
        filename: str,
        progress_callback: typing.Callable = None,
        logger: logging.Logger = None):
    if logger is None:
        logger = logging.getLogger("QGrain")
    else:
        assert isinstance(logger, logging.Logger)
    logger.debug("Start to save artificial dataset.")
    wb = openpyxl.Workbook()
    prepare_styles(wb)
    logger.debug("Creating the `README` sheet.")
    readme_text = \
        """
        It contanins n_components + 3 sheets:
        1. The first sheet is used to put the random settings which were used to generate random parameters.
        2. The second sheet is the generated dataset.
        3. The third sheet stores the random parameters which were used to calulate the component distributions and their mixture.
        4. The left sheets are the distributions of all components.

        Sampling Settings
            Minimum size [μm]: {0}
            Maximum size [μm]: {1}
            Number of Grain Size Classes: {2}
            Precision: {3}
            Noise Decimals: {4}
            Number of Samples: {5}
        """.format(
            dataset.min_μm,
            dataset.max_μm,
            dataset.n_classes,
            dataset.precision,
            dataset.noise,
            dataset.n_samples)
    _write_readme_sheet(wb.active, readme_text)

    def write(row, col, value, style="normal_light"):
        cell = ws.cell(row+1, col+1, value=value)
        cell.style = style

    param_names = list(DISTRIBUTION_CLASS_MAP[dataset.distribution_type].PARAM_NAMES) + ["Weight"]
    n_params = DISTRIBUTION_CLASS_MAP[dataset.distribution_type].N_PARAMS + 1
    if dataset.target is not None:
        logger.debug("Creating the `Random Settings` sheet.")
        ws = wb.create_sheet("Random Settings")
        write(0, 0, "Parameter", style="header")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        for i_param, param_name in enumerate(param_names):
            write(0, i_param*2+1, param_name, style="header")
            ws.merge_cells(start_row=1, start_column=i_param*2+2, end_row=1, end_column=i_param*2+3)
        ws.column_dimensions[column_to_char(0)].width = 16
        for col in range(1, n_params*2+1):
            ws.column_dimensions[column_to_char(col)].width = 16
            if col % 2 == 0:
                write(1, col, "Standard deviation", style="header")
            else:
                write(1, col, "Mean", style="header")
        for row, comp_params in enumerate(dataset.target, 2):
            if row % 2 == 1:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, f"Component{row-1}", style=style)
            for i, (mean, std) in enumerate(comp_params):
                write(row, i*2+1, mean, style=style)
                write(row, i*2+2, std, style=style)
    else:
        logger.warning("Not creating the `Random Settings` sheet.")

    logger.debug("Creating the `Dataset` sheet.")
    ws = wb.create_sheet("Dataset")
    write(0, 0, "Sample Name", style="header")
    ws.column_dimensions[column_to_char(0)].width = 24
    for col, value in enumerate(dataset.classes_μm, 1):
        write(0, col, value, style="header")
        ws.column_dimensions[column_to_char(col)].width = 10
    for i, sample in enumerate(dataset.samples):
        row = i + 1
        if row % 2 == 0:
            style = "normal_dark"
        else:
            style = "normal_light"
        write(row, 0, sample.name, style=style)
        for col, value in enumerate(sample.distribution, 1):
            write(row, col, value, style=style)

        if progress_callback is not None:
            progress_callback(i / dataset.n_samples * 0.2)

    logger.debug("Creating the `Parameters` sheet.")
    ws = wb.create_sheet("Parameters")
    write(0, 0, "Sample Name", style="header")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.column_dimensions[column_to_char(0)].width = 24
    for i in range(dataset.n_components):
        write(0, n_params*i+1, f"Component{i+1}", style="header")
        ws.merge_cells(start_row=1, start_column=n_params*i+2, end_row=1, end_column=n_params*(i+1)+1)
        for j, header_name in enumerate(param_names):
            write(1, n_params*i+1+j, header_name, style="header")
            ws.column_dimensions[column_to_char(n_params*i+1+j)].width = 16
    for i in range(dataset.n_samples):
        row = i + 2
        if row % 2 == 1:
            style = "normal_dark"
        else:
            style = "normal_light"
        write(row, 0, dataset.samples[i].name, style=style)
        for j in range(dataset.n_components):
            for k in range(n_params):
                write(row, n_params*j+k+1, dataset.params[i, k, j], style=style)

        if progress_callback is not None:
            progress_callback(1/dataset.n_samples*0.2 + 0.2)

    for i in range(dataset.n_components):
        logger.debug(f"Creating the `C{i+1}` sheet.")
        ws = wb.create_sheet(f"C{i+1}")
        write(0, 0, "Sample Name", style="header")
        ws.column_dimensions[column_to_char(0)].width = 24
        for col, value in enumerate(dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        for row, sample in enumerate(dataset.samples, 1):
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            for col, value in enumerate(sample.components[i].distribution, 1):
                write(row, col, value, style=style)

        if progress_callback is not None:
            progress_callback(((i*dataset.n_samples + row) / dataset.n_samples*dataset.n_components) * 0.6 + 0.4)
    wb.save(filename)
    wb.close()
    if progress_callback is not None:
        progress_callback(1.0)
    logger.info(f"The artificial dataset has been saved to the Excel file: [{filename}].")


def save_dataset(
        dataset: GrainSizeDataset,
        filename: str,
        progress_callback: typing.Callable = None,
        logger: logging.Logger = None):
    _check_dataset(dataset)
    if logger is None:
        logger = logging.getLogger("QGrain")
    else:
        assert isinstance(logger, logging.Logger)
    logger.debug("Start to save grain size dataset.")
    wb = openpyxl.Workbook()
    prepare_styles(wb)
    logger.debug("Creating the `README` sheet.")
    readme_text = \
        """
        It only contanins one sheet which stores the grain size distributions.
        """
    _write_readme_sheet(wb.active, readme_text)
    logger.debug("Creating the `GSDs` sheet.")
    ws = wb.create_sheet("GSDs")
    _write_dataset_sheet(ws, dataset, progress_callback)
    logger.debug("Saving the workbook to file.")
    wb.save(filename)
    wb.close()
    if progress_callback is not None:
        progress_callback(1.0)
    logger.info(f"The dataset has been saved to the Excel file: [{filename}].")


def save_statistic(
        dataset: GrainSizeDataset,
        filename: str,
        progress_callback: typing.Callable = None,
        logger: logging.Logger = None):
    _check_dataset(dataset)
    if logger is None:
        logger = logging.getLogger("QGrain")
    else:
        assert isinstance(logger, logging.Logger)
    logger.debug("Start to save statistic result.")
    # Calculate
    logger.debug("Calculating the statistic parameters and classification groups of all samples.")
    all_statistics = []
    for i, sample in enumerate(dataset.samples):
        sample_statistics = get_all_statistic(sample.classes_μm, sample.classes_φ, sample.distribution)
        all_statistics.append(sample_statistics)
        if progress_callback is not None:
            progress = (i / dataset.n_samples) * 0.4
            progress_callback(progress)
    wb = openpyxl.Workbook()
    prepare_styles(wb)
    logger.debug("Creating the `README` sheet.")
    readme_text = \
        """
        It contanins 6 sheets.
            1-5. The previous five sheets stores the statistic parameters of five computational methods, respectively.
            6. The last sheet puts the proportions of different size scales and the classification groups.

        The statistic formulas are referred to Blott & Pye (2001)'s work.
        The classification of GSDs is referred to Folk (1954)'s and Blott & Pye (2012)'s scheme.

        References:
            1. Blott, S. J. & Pye, K. Particle size scales and classification of sediment types based on particle size distributions: Review and recommended procedures. Sedimentology 59, 2071–2096 (2012).
            2. Blott, S. J. & Pye, K. GRADISTAT: a grain-size distribution and statistics package for the analysis of unconsolidated sediments. Earth Surf. Process. Landforms 26, 1237–1248 (2001).
            3. Folk, R. L. The Distinction between Grain Size and Mineral Composition in Sedimentary-Rock Nomenclature. The Journal of Geology 62, 344–359 (1954).
        """
    _write_readme_sheet(wb.active, readme_text)
    logger.debug("Creating the `GSDs` sheet.")
    if progress_callback is not None:
        _callback = lambda progress: progress_callback(progress * 0.1 + 0.4)
    else:
        _callback = None
    ws = wb.create_sheet("GSDs")
    _write_dataset_sheet(ws, dataset, progress_callback=_callback)

    def get_keys(method: str):
        small_width = 12
        median_width = 24
        large_width = 48
        if method == "arithmetic":
            unit = "μm"
            keys = [
                (lambda s: s[method]["mean"], f"Mean [{unit}]", small_width),
                (lambda s: s[method]["std"], f"Sorting Coefficient", small_width),
                (lambda s: s[method]["skewness"], f"Skewness", small_width),
                (lambda s: s[method]["kurtosis"], f"Kurtosis", small_width)]
            return keys
        elif method in ("geometric", "logarithmic", "geometric_FW57", "logarithmic_FW57"):
            unit = "μm" if method.startswith("geometric") else "φ"
            keys = [
                (lambda s: s[method]["mean"], f"Mean [{unit}]", small_width),
                (lambda s: s[method]["mean_description"], "Mean Description", median_width),
                (lambda s: s[method]["median"], f"Median [{unit}]", small_width),
                (lambda s: s[method]["mode"], f"Mode [{unit}]", small_width),
                (lambda s: len(s[method]["modes"]), f"Number of Modes", small_width),
                (lambda s: ", ".join([f"{mode: 0.4f}" for mode in s[method]["modes"]]), f"Modes [{unit}]", median_width),
                (lambda s: s[method]["std"], "Sorting Coefficient", small_width),
                (lambda s: s[method]["std_description"], "Sorting Description", median_width),
                (lambda s: s[method]["skewness"], "Skewness", small_width),
                (lambda s: s[method]["skewness_description"], "Skewness Description", median_width),
                (lambda s: s[method]["kurtosis"], "Kurtosis", small_width),
                (lambda s: s[method]["kurtosis_description"], "Kurtosis Description", median_width)]
            return keys
        elif method == "proportion_and_classification":
            keys = [
                (lambda s: ", ".join([f"{p*100:0.4f}" for p in s["GSM_proportion"]]), "(Gravel, Sand, Mud) Proportions [%]", large_width),
                (lambda s: ", ".join([f"{p*100:0.4f}" for p in s["SSC_proportion"]]), "(Sand, Silt, Clay) Proportions [%]", large_width),
                (lambda s: ", ".join([f"{p*100:0.4f}" for p in s["BGSSC_proportion"]]), "(Boulder, Gravel, Sand, Silt, Clay) Proportions [%]", large_width),
                (lambda s: s["group_Folk54"], "Group (Folk, 1954)", median_width),
                (lambda s: s["group_BP12"], "Group (Blott & Pye, 2012)", large_width),
                (lambda s: s["group_BP12_symbol"], "Group Symbol (Blott & Pye, 2012)", median_width)]
            all_scales = _get_all_scales()
            for scale in all_scales:
                func = lambda s, scale=scale: s["proportion"][scale] * 100.0
                name = string.capwords(" ".join(scale)) + " Proportion [%]"
                keys.append((func, name, small_width))
            return keys
        else:
            raise NotImplementedError(method)

    def write(row, col, value, style="normal_light"):
        cell = ws.cell(row+1, col+1, value=value)
        cell.style = style

    methods = ["arithmetic", "geometric", "logarithmic", "geometric_FW57", "logarithmic_FW57", "proportion_and_classification"]
    sheet_names = ["Arithmetic", "Geometric", "Logarithmic", "Geometric_FW57", "Logarithmic_FW57", "Proportion and Classification"]
    for i_method, (method, sheet_name) in enumerate(zip(methods, sheet_names)):
        logger.debug(f"Creating the `{sheet_name}` sheet.")
        ws = wb.create_sheet(sheet_name)
        write(0, 0, "Sample Name", style="header")
        keys = get_keys(method)
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, (func, name, width) in enumerate(keys, 1):
            write(0, col, name, style="header")
            ws.column_dimensions[column_to_char(col)].width = width
        for i_sample, (sample, sample_statistics) in enumerate(zip(dataset.samples, all_statistics)):
            row = i_sample + 1
            if row % 2 == 0:
                style = "normal_dark"
            else:
                style = "normal_light"
            write(row, 0, sample.name, style=style)
            for col, (func, name, width) in enumerate(keys, 1):
                value = func(sample_statistics)
                write(row, col, value, style=style)
            if progress_callback is not None:
                progress = 0.5 + ((i_sample / dataset.n_samples) + i_method) / len(methods) * 0.5
                progress_callback(progress)
    logger.debug("Saving the workbook to file.")
    wb.save(filename)
    wb.close()
    if progress_callback is not None:
        progress_callback(1.0)
    logger.info(f"The statistic result has been saved to the Excel file: [{filename}].")


def save_pca(
        dataset: GrainSizeDataset,
        filename: str,
        progress_callback: typing.Callable = None,
        logger: logging.Logger = None):
    _check_dataset(dataset)
    if logger is None:
        logger = logging.getLogger("QGrain")
    else:
        assert isinstance(logger, logging.Logger)
    logger.debug("Start to save PCA result.")
    # Calculate
    logger.debug("Performing the PCA algorithm to this dataset. The number of PCs is set to 10.")
    pca = PCA(n_components=10)
    transformed = pca.fit_transform(dataset.distribution_matrix)
    components = pca.components_
    ratios = pca.explained_variance_ratio_

    wb = openpyxl.Workbook()
    prepare_styles(wb)
    logger.debug("Creating the `README` sheet.")
    readme_text = \
        """
        It contanins three sheets:
        1. The first sheet is the dataset which was used to perform the PCA algorithm.
        2. The second sheet is used to put the distributions of all PCs.
        3. The third sheet is used to store the PC variations of all samples.

        The base PCA algorithm is implemented by scikit-learn. You can get the details of algorithm from the following website.
        https://scikit-learn.org/stable/modules/generated/sklearn.decomposition.PCA.html
        """
    _write_readme_sheet(wb.active, readme_text)
    logger.debug("Creating the `GSDs` sheet.")
    if progress_callback is not None:
        _callback = lambda progress: progress_callback(progress * 0.2)
    else:
        _callback = None
    ws = wb.create_sheet("GSDs")
    _write_dataset_sheet(ws, dataset, progress_callback=_callback)

    def write(row, col, value, style="normal_light"):
        cell = ws.cell(row+1, col+1, value=value)
        cell.style = style

    logger.debug("Creating the `Distributions of PCs` sheet.")
    ws = wb.create_sheet("Distributions of PCs")
    write(0, 0, "PC", style="header")
    ws.column_dimensions[column_to_char(0)].width = 16
    for col, value in enumerate(dataset.classes_μm, 1):
        write(0, col, value, style="header")
        ws.column_dimensions[column_to_char(col)].width = 10
    for i, component in enumerate(components):
        row = i + 1
        if row % 2 == 0:
            style = "normal_dark"
        else:
            style = "normal_light"
        write(row, 0, f"PC{i+1} ({ratios[i]:0.2%})", style=style)
        for col, value in enumerate(component, 1):
            write(row, col, value, style=style)
    if progress_callback is not None:
        progress_callback(0.3)

    logger.debug("Creating the `Variations of PCs` sheet.")
    ws = wb.create_sheet("Variations of PCs")
    write(0, 0, "Sample Name", style="header")
    ws.column_dimensions[column_to_char(0)].width = 16
    for i in range(10):
        write(0, i+1, f"PC{i+1} ({ratios[i]:0.2%})", style="header")
        ws.column_dimensions[column_to_char(i+1)].width = 10
    for row, varations in enumerate(transformed, 1):
        if row % 2 == 0:
            style = "normal_dark"
        else:
            style = "normal_light"
        write(row, 0, dataset.samples[row-1].name, style=style)
        for col, value in enumerate(varations, 1):
            write(row, col, value, style=style)
        if progress_callback is not None:
            progress_callback(row / dataset.n_samples * 0.7 + 0.3)

    wb.save(filename)
    wb.close()
    if progress_callback is not None:
        progress_callback(1.0)
    logger.info(f"The PCA result has been saved to the Excel file: [{filename}].")


def save_clustering(
        dataset: GrainSizeDataset,
        flags: typing.Iterable[int],
        filename: str,
        progress_callback: typing.Callable = None,
        logger: logging.Logger = None):
    _check_dataset(dataset)
    if logger is None:
        logger = logging.getLogger("QGrain")
    else:
        assert isinstance(logger, logging.Logger)
    logger.debug("Start to save clustering result.")
    flag_set = set(flags)
    n_clusters = len(flag_set)
    typical_samples = [] # type: list[GrainSizeSample]
    temp_flag_set = set()
    for i, flag in enumerate(flags):
        if len(temp_flag_set) == n_clusters:
            break
        if flag not in temp_flag_set:
            typical_samples.append(dataset.samples[i])
            temp_flag_set.add(flag)

    wb = openpyxl.Workbook()
    prepare_styles(wb)
    logger.debug("Creating the `README` sheet.")
    readme_text = \
        """
        It contanins three (or n_clusters + 3) sheets:
        1. The first sheet is the dataset which was used to perform the hierarchy clustering algorithm.
        2. The second sheet is used to put the clustering flags of all samples.
        3. The third sheet is the typical sampels (i.e, the first sample of each cluster was selected).
        4. If the number of clusters less equal to 100, the samples of each cluster will be save to individual sheets.

        The base hierarchy clusrting algorithm is implemented by Scipy. You can get the details of algorithm from the following website.
        https://docs.scipy.org/doc/scipy/reference/cluster.hierarchy.html
        """
    _write_readme_sheet(wb.active, readme_text)
    logger.debug("Creating the `GSDs` sheet.")
    if progress_callback is not None:
        _callback = lambda progress: progress_callback(progress * 0.2)
    else:
        _callback = None
    ws = wb.create_sheet("GSDs")
    _write_dataset_sheet(ws, dataset, progress_callback=_callback)

    def write(row, col, value, style="normal_light"):
        cell = ws.cell(row+1, col+1, value=value)
        cell.style = style

    logger.debug("Creating the `Cluster Flags of Samples` sheet.")
    ws = wb.create_sheet("Cluster Flags of Samples")
    write(0, 0, "Sample Name", style="header")
    write(0, 1, "Cluster Flags", style="header")
    ws.column_dimensions[column_to_char(0)].width = 16
    ws.column_dimensions[column_to_char(1)].width = 16
    for i, (sample, flag) in enumerate(zip(dataset.samples, flags)):
        row = i + 1
        if row % 2 == 0:
            style = "normal_dark"
        else:
            style = "normal_light"
        write(row, 0, sample.name, style=style)
        write(row, 1, flag, style=style)
        if progress_callback is not None:
            if n_clusters <= 100:
                progress_callback(i / dataset.n_samples * 0.1 + 0.2)
            else:
                progress_callback(i / dataset.n_samples * 0.4 + 0.2)

    logger.debug("Creating the `Typical Samples of Clusters` sheet.")
    ws = wb.create_sheet("Typical Samples of Clusters")
    write(0, 0, "Sample Name", style="header")
    ws.column_dimensions[column_to_char(0)].width = 16
    for col, value in enumerate(dataset.classes_μm, 1):
        write(0, col, value, style="header")
        ws.column_dimensions[column_to_char(col)].width = 10
    for i, sample in enumerate(typical_samples):
        row = i + 1
        if row % 2 == 0:
            style = "normal_dark"
        else:
            style = "normal_light"
        write(row, 0, sample.name, style=style)
        for col, value in enumerate(sample.distribution, 1):
            write(row, col, value, style=style)
        if progress_callback is not None:
            if n_clusters <= 100:
                progress_callback(i / dataset.n_samples * 0.1 + 0.3)
            else:
                progress_callback(i / dataset.n_samples * 0.4 + 0.6)

    if n_clusters <= 100:
        for flag in flag_set:
            samples = []
            for sample, in_this_cluster in zip(dataset.samples, np.equal(flags, flag)):
                if in_this_cluster:
                    samples.append(sample)
            logger.debug(f"Creating the `Cluster{flag}` sheet.")
            ws = wb.create_sheet(f"Cluster{flag}")
            write(0, 0, "Sample Name", style="header")
            ws.column_dimensions[column_to_char(0)].width = 16
            for col, value in enumerate(dataset.classes_μm, 1):
                write(0, col, value, style="header")
                ws.column_dimensions[column_to_char(col)].width = 10
            for i, sample in enumerate(samples):
                row = i + 1
                if row % 2 == 0:
                    style = "normal_dark"
                else:
                    style = "normal_light"
                write(row, 0, sample.name, style=style)
                for col, value in enumerate(sample.distribution, 1):
                    write(row, col, value, style=style)
                if progress_callback is not None:
                    progress_callback(i / dataset.n_samples / n_clusters * 0.6 + 0.4)

    wb.save(filename)
    wb.close()
    if progress_callback is not None:
        progress_callback(1.0)
    logger.info(f"The Clustering result has been saved to the Excel file: [{filename}].")


def save_emma(
        result: EMMAResult, filename: str,
        progress_callback: typing.Callable = None,
        logger: logging.Logger = None):
    if logger is None:
        logger = logging.getLogger("QGrain")
    else:
        assert isinstance(logger, logging.Logger)

    logger.debug("Start to save EMMA result.")
    # get the mode size of each end-members
    modes = [(i, result.dataset.classes_μm[np.unravel_index(np.argmax(result.end_members[i]), result.end_members[i].shape)]) for i in range(result.n_members)]
    # sort them by mode size
    modes.sort(key=lambda x: x[1])

    wb = openpyxl.Workbook()
    prepare_styles(wb)
    logger.debug("Creating the `README` sheet.")
    readme_text = \
        """
        It contanins three sheets:
        1. The first sheet is the dataset which was used to perform the EMMA algorithm.
        2. The second sheet is used to put the distributions of all end members.
        3. The third sheet is the proportions of end members of all samples.

        This EMMA algorithm was implemented by QGrian, using the famous machine learning framework, PyTorch.

        EMMA Algorithm Details
            Number of Samples: {0}
            Kernel Type: {1}
            Number of End Members: {2}
            Number of Iterations: {3}
            Spent Time: {4} s

            Computing Device: {5}
            Distance Function: {6}
            Minimum Number of Iterations: {7}
            Maximum Number of Iterations: {8}
            Learning Rate: {9}
            Precision: {10}
            Betas: {11}

        """.format(
            result.dataset.n_samples,
            result.kernel_type.name,
            result.n_members,
            result.n_iterations,
            result.time_spent,
            result.resolver_setting.device,
            result.resolver_setting.distance,
            result.resolver_setting.min_epochs,
            result.resolver_setting.max_epochs,
            result.resolver_setting.learning_rate,
            result.resolver_setting.precision,
            result.resolver_setting.betas)
    _write_readme_sheet(wb.active, readme_text)
    logger.debug("Creating the `GSDs` sheet.")
    if progress_callback is not None:
        _callback = lambda progress: progress_callback(progress * 0.4)
    else:
        _callback = None
    ws = wb.create_sheet("GSDs")
    _write_dataset_sheet(ws, result.dataset, progress_callback=_callback)

    def write(row, col, value, style="normal_light"):
        cell = ws.cell(row+1, col+1, value=value)
        cell.style = style

    logger.debug("Creating the `Distributions of End Members` sheet.")
    ws = wb.create_sheet("Distributions of End Members")
    write(0, 0, "End Member", style="header")
    ws.column_dimensions[column_to_char(0)].width = 16
    for col, value in enumerate(result.dataset.classes_μm, 1):
        write(0, col, value, style="header")
        ws.column_dimensions[column_to_char(col)].width = 10
    for i, (index, _) in enumerate(modes):
        row = i + 1
        if row % 2 == 0:
            style = "normal_dark"
        else:
            style = "normal_light"
        write(row, 0, f"EM{i+1}", style=style)
        for col, value in enumerate(result.end_members[index], 1):
            write(row, col, value, style=style)

    logger.debug("Creating the `Proportions of End Members` sheet.")
    ws = wb.create_sheet("Proportions of End Members")
    write(0, 0, "Sample Name", style="header")
    ws.column_dimensions[column_to_char(0)].width = 16
    for i in range(result.n_members):
        write(0, i+1, f"EM{i+1}", style="header")
        ws.column_dimensions[column_to_char(i+1)].width = 10
    for i, sample_proportions in enumerate(result.proportions):
        row = i + 1
        if row % 2 == 0:
            style = "normal_dark"
        else:
            style = "normal_light"
        write(row, 0, result.dataset.samples[i].name, style=style)
        for col, (index, _) in enumerate(modes, 1):
            write(row, col, sample_proportions[index], style=style)
        if progress_callback is not None:
            progress_callback(i / result.dataset.n_samples * 0.6 + 0.4)

    wb.save(filename)
    wb.close()
    if progress_callback is not None:
        progress_callback(1.0)
    logger.info(f"The EMMA result has been saved to the Excel file: [{filename}].")


def save_ssu(
        results: typing.List[SSUResult], filename: str,
        align_components=False,
        progress_callback: typing.Callable = None,
        logger: logging.Logger = None):
    if logger is None:
        logger = logging.getLogger("QGrain")
    else:
        assert isinstance(logger, logging.Logger)
    logger.debug("Start to save SSU results.")
    # pack the GSDs of samples to the dataset
    dataset = GrainSizeDataset()
    dataset.add_batch(
        results[0].classes_μm,
        [result.sample.name for result in results],
        [result.sample.distribution for result in results],
        need_validation=False)
    max_n_components = max(Counter([result.n_components for result in results]).keys())

    # prepare flags
    flags = []
    if not align_components:
        for result in results:
            flags.extend(range(result.n_components))
    else:
        stacked_components = []
        for result in results:
            for component in result.components:
                stacked_components.append(component.distribution)
        stacked_components = np.array(stacked_components)
        clusers = KMeans(n_clusters=max_n_components)
        flags = clusers.fit_predict(stacked_components)
        # check flags to make it unique
        flag_index = 0
        for result in results:
            result_flags = set()
            for component in result.components:
                if flags[flag_index] in result_flags:
                    if flags[flag_index] == max_n_components:
                        flags[flag_index] = max_n_components-1
                    else:
                        flag_index[flag_index] += 1
                    result_flags.add(flags[flag_index])
                flag_index += 1

        flag_set = set(flags)
        picked = []
        for target_flag in flag_set:
            for i, flag in enumerate(flags):
                if flag == target_flag:
                    picked.append((target_flag, logarithmic(classes_φ, stacked_components[i])["mean"]))
                    break
        picked.sort(key=lambda x: x[1])
        flag_map = {flag: index for index, (flag, _) in enumerate(picked)}
        flags = np.array([flag_map[flag] for flag in flags])

    wb = openpyxl.Workbook()
    prepare_styles(wb)
    logger.debug("Creating the `README` sheet.")
    readme_text = \
        """
        It contanins 4 + max number of components sheets:
        1. The first sheet is used to put the grain size distributions of corresponding samples.
        2. The second sheet is used to put the fitting information and resolved parameters of SSU results.
        3. The third sheet stores the statistic parameters of each components.
        4. The fouth sheet is used to put the distributions of unmixed components and the sum.
        5. Other sheets severally store the distributions of each component group.

        The SSU algorithm is implemented by QGrain.

        """
    _write_readme_sheet(wb.active, readme_text)
    logger.debug("Creating the `GSDs` sheet.")
    if progress_callback is not None:
        _callback = lambda progress: progress_callback(progress * 0.1)
    else:
        _callback = None
    ws = wb.create_sheet("GSDs")
    _write_dataset_sheet(ws, dataset, progress_callback=_callback)

    def write(row, col, value, style="normal_light"):
        cell = ws.cell(row+1, col+1, value=value)
        cell.style = style

    logger.debug("Creating the `Information of Fitting` sheet.")
    ws = wb.create_sheet("Information of Fitting")
    write(0, 0, "Sample Name", style="header")
    ws.column_dimensions[column_to_char(0)].width = 16
    headers = ["Distribution Type",
               "Number of Components",
               "Resolver Settings",
               "Initial Guess",
               "Resolved Parameters",
               "Spent Time [s]",
               "Number of Iterations",
               "Final Distance [log10MSE]"]
    for col, value in enumerate(headers, 1):
        write(0, col, value, style="header")
        if col in (3, 4, 5):
            ws.column_dimensions[column_to_char(col)].width = 32
        else:
            ws.column_dimensions[column_to_char(col)].width = 10
    for i, result in enumerate(results):
        row = i + 1
        if row % 2 == 0:
            style = "normal_dark"
        else:
            style = "normal_light"
        write(row, 0, result.sample.name, style=style)
        write(row, 1, result.distribution_type.name, style=style)
        write(row, 2, result.n_components, style=style)
        write(row, 3, "Default" if result.task.resolver_setting is None else result.task.resolver_setting.__str__(), style=style)
        write(row, 4, "None" if result.task.initial_guess is None else result.task.initial_guess.__str__(), style=style)
        write(row, 5, result.func_args.__str__(), style=style)
        write(row, 6, result.time_spent, style=style)
        write(row, 7, result.n_iterations, style=style)
        write(row, 8, result.get_distance("log10MSE"), style=style)
        if progress_callback is not None:
            progress_callback(i / dataset.n_samples * 0.1 + 0.1)

    logger.debug("Creating the `Statistic Moments` sheet.")
    ws = wb.create_sheet("Statistic Moments")
    write(0, 0, "Sample Name", style="header")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.column_dimensions[column_to_char(0)].width = 16
    headers = []
    sub_headers = ["Proportion",
                   "Mean [φ]",
                   "Mean [μm]",
                   "Standard Deviation [φ]",
                   "Standard Deviation [μm]",
                   "Skewness",
                   "Kurtosis"]
    for i in range(max_n_components):
        write(0, i*len(sub_headers)+1, f"C{i+1}", style="header")
        ws.merge_cells(start_row=1, start_column=i*len(sub_headers)+2, end_row=1, end_column=(i+1)*len(sub_headers)+1)
        headers.extend(sub_headers)
    for col, value in enumerate(headers, 1):
        write(1, col, value, style="header")
        ws.column_dimensions[column_to_char(col)].width = 10
    flag_index = 0
    for i, result in enumerate(results):
        row = i + 2
        if row % 2 == 0:
            style = "normal_light"
        else:
            style = "normal_dark"
        write(row, 0, result.sample.name, style=style)
        for component in result.components:
            index = flags[flag_index]
            s = get_all_statistic(result.classes_μm, result.classes_φ, component.distribution)
            write(row, index*len(sub_headers)+1, component.proportion, style=style)
            write(row, index*len(sub_headers)+2, s["logarithmic"]["mean"], style=style)
            write(row, index*len(sub_headers)+3, s["geometric"]["mean"], style=style)
            write(row, index*len(sub_headers)+4, s["logarithmic"]["std"], style=style)
            write(row, index*len(sub_headers)+5, s["geometric"]["std"], style=style)
            write(row, index*len(sub_headers)+6, s["logarithmic"]["skewness"], style=style)
            write(row, index*len(sub_headers)+7, s["logarithmic"]["kurtosis"], style=style)
            flag_index += 1
        if progress_callback is not None:
            progress_callback(i / dataset.n_samples * 0.1 + 0.2)

    logger.debug("Creating the `Unmixed Components` sheet.")
    ws = wb.create_sheet("Unmixed Components")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    write(0, 0, "Sample Name", style="header")
    ws.column_dimensions[column_to_char(0)].width = 16
    for col, value in enumerate(dataset.classes_μm, 2):
        write(0, col, value, style="header")
        ws.column_dimensions[column_to_char(col)].width = 10
    row = 1
    for result_index, result in enumerate(results):
        if result_index % 2 == 0:
            style = "normal_light"
        else:
            style = "normal_dark"
        write(row, 0, result.sample.name, style=style)
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+result.n_components+1, end_column=1)
        for component_i, component in enumerate(result.components, 1):
            write(row, 1, f"C{component_i}", style=style)
            for col, value in enumerate(component.distribution*component.proportion, 2):
                write(row, col, value, style=style)
            row += 1
        write(row, 1, "Sum", style=style)
        for col, value in enumerate(result.distribution, 2):
            write(row, col, value, style=style)
        row += 1
        if progress_callback is not None:
            progress_callback(i / dataset.n_samples * 0.2 + 0.3)

    logger.debug("Creating separate sheets for all components.")
    ws_dict = {}
    flag_set = set(flags)
    for flag in flag_set:
        ws = wb.create_sheet(f"C{flag+1}")
        write(0, 0, "Sample Name", style="header")
        ws.column_dimensions[column_to_char(0)].width = 16
        for col, value in enumerate(dataset.classes_μm, 1):
            write(0, col, value, style="header")
            ws.column_dimensions[column_to_char(col)].width = 10
        ws_dict[flag] = ws

    flag_index = 0
    for i, result in enumerate(results):
        row = i + 1
        if row % 2 == 0:
            style = "normal_dark"
        else:
            style = "normal_light"

        for component in result.components:
            flag = flags[flag_index]
            ws = ws_dict[flag]
            write(row, 0, result.sample.name, style=style)
            for col, value in enumerate(component.distribution, 1):
                write(row, col, value, style=style)
            flag_index += 1
        if progress_callback is not None:
            progress_callback(i / dataset.n_samples * 0.5 + 0.5)

    wb.save(filename)
    wb.close()
    if progress_callback is not None:
        progress_callback(1.0)
    logger.info(f"The SSU results have been saved to the Excel file: [{filename}].")
